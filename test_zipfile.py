import csv
import os
import zipfile
from os.path import basename

from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook

path_from = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')  # путь к папке "resources"
path_to = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'archive')  # путь к папке "archive"
path_to_zip = os.path.join(path_to, 'zip_doc.zip')  # путь к "zip_doc"


# создаем архив zip и проверяем количество файлов в нем
def test_new_archive():
    with zipfile.ZipFile(path_to_zip, 'w') as zip_f:
        for file in os.listdir(path_from):
            add_file = os.path.join(path_from, file)
            zip_f.write(add_file, basename(add_file))
    file_list = zip_f.namelist()
    assert len(file_list) == 3


# читаем csv файл из архива и проверяем текст

def test_read_csv():
    with zipfile.ZipFile(path_to_zip) as zip_f:
        file = zip_f.extract('ab_test_groups.csv')
    with open(file) as csv_file:
        csv_file = csv.reader(csv_file)
        csv_list = []
        for i in csv_file:
            text = " ".join(i).replace(";", " ")
            csv_list.append(text)

        assert csv_list[10] == '10 B 132'

    os.remove(file)


# читаем xlsx файл из архива и проверяем значение

def test_read_xls():
    with zipfile.ZipFile(path_to_zip) as zip_f:
        file = zip_f.extract('шаблон поставки.xlsx')
        book = load_workbook(file)
        sheet = book.active
        check_value = sheet.cell(row=7, column=2).value
        assert check_value == 'GT168'

    os.remove(file)


# читаем PDF файл из архива и проверяем количество страниц

def test_read_pdf():
    with zipfile.ZipFile(path_to_zip) as zip_f:
        file = zip_f.extract('ChildFund-Sponsor-Visit.pdf')
        reader = PdfReader(file)

        assert len(reader.pages) == 2
    os.remove(file)
